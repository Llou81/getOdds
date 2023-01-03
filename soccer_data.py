import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ID = [8, 158, 9, 10, 11, 12, 34, 156, 35, 151, 36, 88, 153, 161, 913, 37]


def main():
    matches = getMatches(ID)
    updateFile("Odds.xlsx", matches)


def getData(url):
    response = requests.request("GET", url)
    data = response.json()

    return data


def updateFile(path, lista):
    wb = load_workbook(path)
    ws = wb.active
    row = ws.max_row + 1
    for item in lista:
        item.pop(0)
        i = 1
        for odd in item:
            char = get_column_letter(i)
            ws[f"{char}{row}"] = odd
            i += 1
        row += 1
    wb.save(path)
    print("File is successfully updated")


def getMatches(list):
    matches = []

    for _ in list:

        url = f"https://soccerbet.rs/api/Prematch/GetCompetitionMatches?competitionId={_}&timeFrameOption=4"
        data = getData(url)

        for match in data:
            matches.append(
                [match["Id"], match["HomeCompetitorName"], match["AwayCompetitorName"]]
            )

    for _ in range(0, len(matches)):

        url = f"https://soccerbet.rs/api/Prematch/GetMatchBets?matchId={matches[_][0]}"
        data = getData(url)
        for kvota in data:
            matches[_].append(kvota["Odds"])

    return matches


if __name__ == "__main__":
    main()
